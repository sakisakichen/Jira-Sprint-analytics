import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
import re

# ================= CONFIG =================
INPUT_FILE = "jira_query_all.xlsx"
OUTPUT_FILE = "jira_dashboard.xlsx"
THRESHOLD = 640

# Teams to include (T numbers)
INCLUDED_T_NUMS = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46}

# ================= STYLES =================
DARK_BLUE    = "1F3864"
MED_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
RED          = "FF0000"
LIGHT_RED    = "FFE5E5"
ORANGE       = "FF6600"
LIGHT_ORANGE = "FFF0E5"
YELLOW       = "FFD700"
LIGHT_YELLOW = "FFFDE5"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F2F2F2"
DARK_GRAY    = "404040"

def header_style(cell, bg=DARK_BLUE, fg=WHITE, size=13):
    cell.font = Font(name="Arial", bold=True, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def subheader_style(cell, bg=MED_BLUE, fg=WHITE):
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def border_cell(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def extract_sprint_num(sprint_name):
    """Extract T number - use the LAST T number found in sprint name"""
    if pd.isna(sprint_name):
        return 9999
    matches = re.findall(r'T(\d+)', str(sprint_name))
    return int(matches[-1]) if matches else 9999

def sprint_label(sprint_name):
    """Convert 'Sprint CS122 C2 T5' -> 'T5' for chart label"""
    if pd.isna(sprint_name):
        return str(sprint_name)
    matches = re.findall(r'T(\d+)', str(sprint_name))
    return f"T{matches[-1]}" if matches else str(sprint_name)

# ================= LOAD DATA =================
print("Loading data...")
df = pd.read_excel(INPUT_FILE, sheet_name="All_Queries")
df.columns = df.columns.str.strip()
df["Sigma Time Spent (hrs)"] = pd.to_numeric(df["Sigma Time Spent (hrs)"], errors="coerce")
df["Story Points"] = pd.to_numeric(df["Story Points"], errors="coerce")
df["T_num"] = df["Sprint"].apply(extract_sprint_num)

# Filter to included teams only
df_filtered = df[df["T_num"].isin(INCLUDED_T_NUMS)].copy()

# ================= COMPUTE SUMMARIES =================

# --- Sprint Hours (filtered, sorted by T number) ---
sprint_hours = (
    df_filtered.groupby("Sprint")["Sigma Time Spent (hrs)"]
    .sum().reset_index()
    .rename(columns={"Sigma Time Spent (hrs)": "Total Hours"})
)
story_hours = (
    df_filtered[df_filtered["Issue Type"].str.lower() == "story"]
    .groupby("Sprint")["Sigma Time Spent (hrs)"]
    .sum().reset_index()
    .rename(columns={"Sigma Time Spent (hrs)": "Story Hours"})
)
sprint_hours = sprint_hours.merge(story_hours, on="Sprint", how="left")
sprint_hours["Story Hours"] = sprint_hours["Story Hours"].fillna(0).round(0).astype(int)
sprint_hours["Total Hours"] = sprint_hours["Total Hours"].round(0).astype(int)
sprint_hours["T_num"] = sprint_hours["Sprint"].apply(extract_sprint_num)
sprint_hours["T_label"] = sprint_hours["Sprint"].apply(sprint_label)
sprint_hours = sprint_hours.sort_values("T_num").reset_index(drop=True)
sprint_hours["Gap to Standard"] = sprint_hours["Story Hours"].apply(lambda x: max(0, THRESHOLD - x))

# --- Missing Story Points ---
stories = df_filtered[df_filtered["Issue Type"].str.lower() == "story"].copy()
missing_sp = stories[stories["Story Points"].isna()].copy()
missing_sp = missing_sp.sort_values("T_num").reset_index(drop=True)
missing_sp_by_sprint = (
    missing_sp.groupby("Sprint").size().reset_index(name="Missing Count")
)
missing_sp_by_sprint["T_num"] = missing_sp_by_sprint["Sprint"].apply(extract_sprint_num)
missing_sp_by_sprint = missing_sp_by_sprint.sort_values("T_num").reset_index(drop=True)

# --- No Assignee OR No Hours, only Story issue type ---
df_stories = df_filtered[df_filtered["Issue Type"].str.lower() == "story"].copy()
no_assignee = df_stories["Assignee"].isna()
no_hours = df_stories["Sigma Time Spent (hrs)"].isna() | (df_stories["Sigma Time Spent (hrs)"] == 0)
flagged = df_stories[no_assignee | no_hours].copy()
flagged["No Assignee"] = flagged["Assignee"].isna()
flagged["No Hours"]    = flagged["Sigma Time Spent (hrs)"].isna() | (flagged["Sigma Time Spent (hrs)"] == 0)
flagged["Issue Flag"]  = flagged.apply(
    lambda r: "No Assignee & No Hours" if r["No Assignee"] and r["No Hours"]
    else ("No Assignee" if r["No Assignee"] else "No Hours"),
    axis=1
)
flagged = flagged.sort_values(["T_num", "Issue Flag"]).reset_index(drop=True)

# ================= BUILD WORKBOOK =================
wb = Workbook()

# ======================================================
# SHEET 1: SUMMARY
# ======================================================
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 45

ws1.merge_cells("A1:F1")
ws1["A1"] = "JIRA SPRINT DASHBOARD — CS122 C2"
header_style(ws1["A1"], size=14)

ws1.row_dimensions[3].height = 25
ws1.row_dimensions[4].height = 35

stories_only = df_filtered[df_filtered["Issue Type"].str.lower() == "story"]
kpis = [
    ("Total Stories",              len(stories_only),                                                   DARK_GRAY),
    ("Story Hours Logged",         f"{stories_only['Sigma Time Spent (hrs)'].sum():,.0f}",              DARK_GRAY),
    ("Sprints",                    df_filtered["Sprint"].dropna().nunique(),                             DARK_GRAY),
    ("Sprints Below 640 Story Hrs",len(sprint_hours[sprint_hours["Story Hours"] < THRESHOLD]),           RED),
    ("Missing Story Points",       len(missing_sp),                                                      ORANGE),
    ("No Assignee or No Hrs",      len(flagged),                                                         RED),
]

for i, (label, value, fg) in enumerate(kpis):
    col = i + 1
    set_col_width(ws1, col, 22)
    lc = ws1.cell(row=3, column=col, value=label)
    subheader_style(lc)
    border_cell(lc)
    vc = ws1.cell(row=4, column=col, value=value)
    vc.font = Font(name="Arial", bold=True, size=16, color=fg)
    vc.fill = PatternFill("solid", start_color=LIGHT_BLUE)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    border_cell(vc)

# ======================================================
# SHEET 2: HOURS PER SPRINT
# ======================================================
ws2 = wb.create_sheet("01 Hours per Sprint")
ws2.sheet_view.showGridLines = False
ws2.row_dimensions[1].height = 40

ws2.merge_cells("A1:D1")
ws2["A1"] = "HOURS PER SPRINT"
header_style(ws2["A1"])

ws2.merge_cells("A2:D2")
ws2["A2"] = f"Standard Threshold: {THRESHOLD} hrs  |  Sprints below {THRESHOLD} hrs are highlighted in red"
ws2["A2"].font = Font(name="Arial", italic=True, color=DARK_GRAY, size=10)
ws2["A2"].alignment = Alignment(horizontal="center")

# Headers
headers = ["Sprint", "T Label", "Story Hours", "Total Hours", "Gap to Standard (hrs)"]
widths  = [35, 10, 18, 18, 25]
for col, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws2.cell(row=4, column=col, value=h)
    subheader_style(c)
    set_col_width(ws2, col, w)
    border_cell(c)

for row_idx, row in enumerate(sprint_hours.itertuples(), 5):
    is_below = row._3 < THRESHOLD  # Story Hours is col 3
    bg = LIGHT_RED if is_below else (LIGHT_GRAY if row_idx % 2 == 0 else WHITE)
    gap = int(max(0, THRESHOLD - row._3))
    vals = [row.Sprint, row.T_label, row._3, row._4, gap if is_below else ""]
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", color=RED if is_below else DARK_GRAY, size=10,
                      bold=is_below and col in [1, 2, 3])
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        border_cell(c)

# Bar chart — use T_label as category, blue bars only, red 640 line
last_data_row = 4 + len(sprint_hours)

# Write T_label to col 6 for chart categories
ws2.cell(row=4, column=6, value="T Label (chart)")
for i, row in enumerate(sprint_hours.itertuples(), 5):
    ws2.cell(row=i, column=6, value=row.T_label)

# Write threshold values to col 7
ws2.cell(row=4, column=7, value="Standard (640)")
for i in range(5, last_data_row + 1):
    ws2.cell(row=i, column=7, value=THRESHOLD)

from openpyxl.chart import LineChart

# Bar chart - Story Hours only
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Story Hours per Sprint"
chart1.y_axis.title = "Hours"
chart1.x_axis.title = "Team Sprint"
chart1.height = 15
chart1.width = 32
chart1.y_axis.numFmt = '#,##0'

story_ref = Reference(ws2, min_col=3, min_row=4, max_row=last_data_row)
cats_ref = Reference(ws2, min_col=6, min_row=5, max_row=last_data_row)
chart1.add_data(story_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.solidFill = "2E75B6"

# Data labels on bars
chart1.series[0].dLbls = DataLabelList()
chart1.series[0].dLbls.showVal = True
chart1.series[0].dLbls.showSerName = False
chart1.series[0].dLbls.showCatName = True
chart1.series[0].dLbls.position = "ctr"

# Line chart for 640 threshold
line_chart = LineChart()
line_ref = Reference(ws2, min_col=7, min_row=4, max_row=last_data_row)
line_chart.add_data(line_ref, titles_from_data=True)
line_chart.series[0].graphicalProperties.line.solidFill = "FF0000"
line_chart.series[0].graphicalProperties.line.width = 20000
line_chart.series[0].smooth = False

# Combine bar + line
chart1 += line_chart
ws2.add_chart(chart1, f"A{last_data_row + 2}")

# ======================================================
# SHEET 3: MISSING STORY POINTS
# ======================================================
ws3 = wb.create_sheet("02 Missing Story Points")
ws3.sheet_view.showGridLines = False
ws3.row_dimensions[1].height = 40

ws3.merge_cells("A1:F1")
ws3["A1"] = "STORIES WITHOUT STORY POINTS"
header_style(ws3["A1"])

total_stories = len(stories)
missing_count = len(missing_sp)
pct = (missing_count / total_stories * 100) if total_stories > 0 else 0
ws3.merge_cells("A2:F2")
ws3["A2"] = f"Total Stories: {total_stories}  |  Missing Story Points: {missing_count}  |  {pct:.1f}% incomplete"
ws3["A2"].font = Font(name="Arial", italic=True, color=DARK_GRAY, size=10)
ws3["A2"].alignment = Alignment(horizontal="center")

for col, (h, w) in enumerate(zip(["Sprint", "Missing Count"], [35, 18]), 1):
    c = ws3.cell(row=4, column=col, value=h)
    subheader_style(c)
    set_col_width(ws3, col, w)
    border_cell(c)

for row_idx, row in enumerate(missing_sp_by_sprint.itertuples(), 5):
    bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
    for col, val in enumerate([row.Sprint, row._2], 1):
        c = ws3.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", size=10, color=DARK_GRAY)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        border_cell(c)

last_summary_row = 4 + len(missing_sp_by_sprint)

pie = PieChart()
pie.title = "Missing Story Points by Sprint"
pie.height = 12
pie.width = 18
pie_data = Reference(ws3, min_col=2, min_row=4, max_row=last_summary_row)
pie_cats = Reference(ws3, min_col=1, min_row=5, max_row=last_summary_row)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
ws3.add_chart(pie, "D4")

detail_start = last_summary_row + 3
ws3.merge_cells(f"A{detail_start}:F{detail_start}")
ws3[f"A{detail_start}"] = "Detail: Stories Missing Story Points (sorted by Sprint)"
subheader_style(ws3[f"A{detail_start}"], bg=DARK_BLUE)

detail_headers = ["Issue Key", "Summary", "Sprint", "Status", "Assignee"]
widths = [18, 50, 35, 20, 25]
for col, (h, w) in enumerate(zip(detail_headers, widths), 1):
    c = ws3.cell(row=detail_start + 1, column=col, value=h)
    subheader_style(c)
    set_col_width(ws3, col, w)
    border_cell(c)

for row_idx, row in enumerate(missing_sp[["Issue Key", "Summary", "Sprint", "Status", "Assignee"]].itertuples(index=False), detail_start + 2):
    bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", size=10, color=DARK_GRAY)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        border_cell(c)

# ======================================================
# SHEET 4: NO ASSIGNEE OR NO HOURS
# ======================================================
ws4 = wb.create_sheet("03 No Assignee or No Hours")
ws4.sheet_view.showGridLines = False
ws4.row_dimensions[1].height = 40

ws4.merge_cells("A1:H1")
ws4["A1"] = "ISSUES WITH NO ASSIGNEE OR NO HOURS LOGGED (Excluding Sprint Detail)"
header_style(ws4["A1"])

ws4.merge_cells("A2:H2")
ws4["A2"] = f"Total Flagged: {len(flagged)}  |  Sprints Affected: {flagged['Sprint'].nunique()}  |  Sprint Detail issues excluded"
ws4["A2"].font = Font(name="Arial", italic=True, color=RED, size=10)
ws4["A2"].alignment = Alignment(horizontal="center")

# Summary by Sprint
summary_by_sprint = (
    flagged.groupby(["Sprint", "Issue Flag"]).size()
    .reset_index(name="Count")
)
summary_by_sprint["T_num"] = summary_by_sprint["Sprint"].apply(extract_sprint_num)
summary_by_sprint = summary_by_sprint.sort_values(["T_num", "Issue Flag"]).reset_index(drop=True)

for col, (h, w) in enumerate(zip(["Sprint", "Issue Flag", "Count"], [35, 30, 12]), 1):
    c = ws4.cell(row=4, column=col, value=h)
    subheader_style(c)
    set_col_width(ws4, col, w)
    border_cell(c)

for row_idx, row in enumerate(summary_by_sprint.itertuples(), 5):
    is_both      = row._2 == "No Assignee & No Hours"
    is_no_assign = row._2 == "No Assignee"
    bg = LIGHT_RED if is_both else (LIGHT_ORANGE if is_no_assign else LIGHT_YELLOW)
    fg = RED if is_both else (ORANGE if is_no_assign else "996600")
    for col, val in enumerate([row.Sprint, row._2, row.Count], 1):
        c = ws4.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", size=10, color=fg, bold=is_both)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        border_cell(c)

last_summary_row4 = 4 + len(summary_by_sprint)

# Legend
legend_row = last_summary_row4 + 1
for col, (label, bg, fg) in enumerate([
    ("No Assignee & No Hours", LIGHT_RED,    RED),
    ("No Assignee only",       LIGHT_ORANGE, ORANGE),
    ("No Hours only",          LIGHT_YELLOW, "996600"),
], 1):
    c = ws4.cell(row=legend_row, column=col, value=label)
    c.font = Font(name="Arial", size=9, color=fg, bold=True)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center")
    border_cell(c)

# Detail table
detail_start4 = legend_row + 2
ws4.merge_cells(f"A{detail_start4}:H{detail_start4}")
ws4[f"A{detail_start4}"] = "Detail: All Flagged Issues (sorted by Sprint T number)"
subheader_style(ws4[f"A{detail_start4}"], bg=DARK_BLUE)

detail_headers4 = ["Issue Key", "Issue Type", "SR Number", "Summary", "Sprint", "Status", "Assignee", "Issue Flag"]
widths4         = [18,          20,            18,           50,        35,       20,        25,          28]
for col, (h, w) in enumerate(zip(detail_headers4, widths4), 1):
    c = ws4.cell(row=detail_start4 + 1, column=col, value=h)
    subheader_style(c)
    set_col_width(ws4, col, w)
    border_cell(c)

detail_cols = ["Issue Key", "Issue Type", "SR Number", "Summary", "Sprint", "Status", "Assignee", "Issue Flag"]
for row_idx, row in enumerate(flagged[detail_cols].itertuples(index=False), detail_start4 + 2):
    flag         = row[-1]
    is_both      = flag == "No Assignee & No Hours"
    is_no_assign = flag == "No Assignee"
    bg = LIGHT_RED if is_both else (LIGHT_ORANGE if is_no_assign else LIGHT_YELLOW)
    fg = RED if is_both else (ORANGE if is_no_assign else "996600")
    for col, val in enumerate(row, 1):
        c = ws4.cell(row=row_idx, column=col, value=val)
        c.font = Font(name="Arial", size=10,
                      color=fg if col == 8 else DARK_GRAY,
                      bold=col == 8 and is_both)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        border_cell(c)

# ================= SAVE =================
wb.save(OUTPUT_FILE)
print(f"\n Dashboard exported → {OUTPUT_FILE}")
print(f"  Total issues (filtered): {len(df_filtered)}")
print(f"  Flagged (no assign/hrs): {len(flagged)}")
print(f"  Missing story points:    {len(missing_sp)}")
print(f"  Sprints below 640hrs:    {len(sprint_hours[sprint_hours['Total Hours'] < THRESHOLD])}")
